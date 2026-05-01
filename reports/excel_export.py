import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from analytics.stats import calculate_class_stats, calculate_student_stats
from database import get_connection
import pandas as pd

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'reports')

def generate_excel_report():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    class_stats = calculate_class_stats()
    
    conn = get_connection()
    df = pd.read_sql_query('''
        SELECT s.name, s.roll_number, s.department, s.semester,
               r.subject, r.marks, r.max_marks
        FROM students s
        JOIN results r ON s.id = r.student_id
    ''', conn)
    conn.close()
    
    wb = openpyxl.Workbook()
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2980B9")
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws1 = wb.active
    ws1.title = "Leaderboard"
    
    ws1.merge_cells('A1:F1')
    ws1['A1'] = 'Class Leaderboard & Analytics Summary'
    ws1['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws1['A1'].fill = PatternFill("solid", fgColor="1A5276")
    ws1['A1'].alignment = center
    ws1.row_dimensions[1].height = 30
    
    ws1['A3'] = 'Class Statistics'
    ws1['A3'].font = Font(bold=True, size=12)
    
    stats_data = [
        ('Total Students', class_stats['total_students']),
        ('Class Average', f"{class_stats['class_average']}%"),
        ('Highest Score', f"{class_stats['highest_score']}%"),
        ('Lowest Score', f"{class_stats['lowest_score']}%"),
        ('Median Score', f"{class_stats['median']}%"),
        ('Std Deviation', f"{class_stats['std_deviation']}%"),
        ('Pass Rate', f"{class_stats['pass_rate']}%"),
    ]
    
    for i, (label, value) in enumerate(stats_data, start=4):
        ws1[f'A{i}'] = label
        ws1[f'A{i}'].font = Font(bold=True)
        ws1[f'B{i}'] = value
        ws1[f'B{i}'].alignment = center
    
    ws1['D3'] = 'Leaderboard'
    ws1['D3'].font = Font(bold=True, size=12)
    
    headers = ['Rank', 'Name', 'Percentage', 'Percentile']
    for col, header in enumerate(headers, start=4):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    
    grade_colors = {'O': 'E8F8F5', 'A+': 'EBF5FB', 'A': 'F4ECF7',
                    'B+': 'FEF9E7', 'B': 'FDFEFE', 'C': 'FEF5E7', 'F': 'FDEDEC'}
    
    for row_idx, student in enumerate(class_stats['leaderboard'], start=5):
        student_stats = calculate_student_stats(student['student_id'])
        grade = student_stats['grade'] if student_stats else 'N/A'
        fill_color = grade_colors.get(grade, 'FFFFFF')
        row_fill = PatternFill("solid", fgColor=fill_color)
        
        row_data = [student['rank'], student['name'],
                    f"{student['percentage']}%", f"{student['percentile']}%"]
        
        for col_idx, value in enumerate(row_data, start=4):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = center
            cell.border = thin
    
    ws2 = wb.create_sheet("Raw Results")
    
    ws2_headers = ['Name', 'Roll Number', 'Department', 'Semester', 'Subject', 'Marks', 'Max Marks', 'Percentage']
    for col, header in enumerate(ws2_headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        pct = round((row['marks'] / row['max_marks']) * 100, 2)
        row_data = [row['name'], row['roll_number'], row['department'],
                    row['semester'], row['subject'], row['marks'], row['max_marks'], f"{pct}%"]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center
            cell.border = thin
    
    ws3 = wb.create_sheet("Subject Averages")
    ws3['A1'] = 'Subject'
    ws3['A1'].font = header_font
    ws3['A1'].fill = header_fill
    ws3['B1'] = 'Average %'
    ws3['B1'].font = header_font
    ws3['B1'].fill = header_fill
    
    for row_idx, (subject, avg) in enumerate(class_stats['subject_averages'].items(), start=2):
        ws3[f'A{row_idx}'] = subject
        ws3[f'B{row_idx}'] = f"{avg}%"
        ws3[f'B{row_idx}'].alignment = center
    
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4
    
    filepath = os.path.join(OUTPUT_DIR, 'analytics_report.xlsx')
    wb.save(filepath)
    return {"success": True, "file": filepath, "filename": "analytics_report.xlsx"}